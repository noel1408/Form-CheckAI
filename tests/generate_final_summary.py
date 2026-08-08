import os
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def generate_summary(search_dirs=["artifacts", "Test Results/Excel"]):
    # 1. Find all Excel reports
    report_files = []
    for d in search_dirs:
        if os.path.exists(d):
            # Find all xlsx files recursively, ignoring the final summary if it already exists
            found = glob.glob(os.path.join(d, '**/*.xlsx'), recursive=True)
            for f in found:
                if "Final_Test_Summary.xlsx" not in f:
                    report_files.append(f)

    # 2. Extract Data
    job_metrics = {}
    failed_tests = []
    
    total_tests = 0
    total_passed = 0
    total_failed = 0
    total_skipped = 0
    total_errors = 0
    total_duration = 0.0

    for file in report_files:
        try:
            wb = load_workbook(file, data_only=True)
            if "Test Results" not in wb.sheetnames:
                continue
            ws = wb["Test Results"]
            
            # Find column indices
            header = {cell.value: idx for idx, cell in enumerate(ws[1])}
            
            suite_col = header.get("Test Suite", 1)
            name_col = header.get("Test Name", 2)
            status_col = header.get("Status", 3)
            duration_col = header.get("Duration (s)", 4)
            error_col = header.get("Error Message", 5)

            current_job = "Unknown Job"
            job_duration = 0.0
            job_passed = 0
            job_failed = 0
            job_skipped = 0
            job_errors = 0

            # Process rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[suite_col]:
                    continue
                current_job = str(row[suite_col])
                test_name = str(row[name_col]) if row[name_col] else ""
                status = str(row[status_col]).upper() if row[status_col] else "UNKNOWN"
                
                try:
                    duration = float(row[duration_col]) if row[duration_col] else 0.0
                except (ValueError, TypeError):
                    duration = 0.0
                    
                error_msg = str(row[error_col]) if row[error_col] else ""

                if current_job not in job_metrics:
                    job_metrics[current_job] = {
                        "Total": 0, "Passed": 0, "Failed": 0, "Skipped": 0, "Errors": 0, "Duration": 0.0
                    }
                
                job_metrics[current_job]["Total"] += 1
                job_metrics[current_job]["Duration"] += duration
                total_duration += duration
                total_tests += 1

                if status == "PASSED":
                    job_metrics[current_job]["Passed"] += 1
                    total_passed += 1
                elif status == "FAILED":
                    job_metrics[current_job]["Failed"] += 1
                    total_failed += 1
                    failed_tests.append({
                        "Job": current_job,
                        "Test": test_name,
                        "Error": error_msg,
                        "Duration": duration
                    })
                elif status == "SKIPPED":
                    job_metrics[current_job]["Skipped"] += 1
                    total_skipped += 1
                else:
                    job_metrics[current_job]["Errors"] += 1
                    total_errors += 1
                    failed_tests.append({
                        "Job": current_job,
                        "Test": test_name,
                        "Error": error_msg,
                        "Duration": duration
                    })
        except Exception as e:
            print(f"Error reading {file}: {e}")

    # 3. Create Final Workbook
    out_wb = Workbook()
    
    # --- Sheet 1: Overall Summary ---
    ws1 = out_wb.active
    ws1.title = "Overall Summary"
    
    ws1.append(["Metric", "Value"])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill

    overall_pass_pct = (total_passed / total_tests * 100) if total_tests > 0 else 0.0
    overall_fail_pct = (total_failed / total_tests * 100) if total_tests > 0 else 0.0
    
    jobs_passed = sum(1 for j in job_metrics.values() if j["Failed"] == 0 and j["Errors"] == 0)
    jobs_failed = sum(1 for j in job_metrics.values() if j["Failed"] > 0 or j["Errors"] > 0)

    summary_data = [
        ["Total Test Jobs", len(job_metrics)],
        ["Total Tests", total_tests],
        ["Total Passed", total_passed],
        ["Total Failed", total_failed],
        ["Total Skipped", total_skipped],
        ["Total Errors", total_errors],
        ["Overall Pass %", f"{overall_pass_pct:.2f}%"],
        ["Overall Fail %", f"{overall_fail_pct:.2f}%"],
        ["Total Pipeline/Test Duration (s)", f"{total_duration:.2f}"],
        ["Jobs Passed", jobs_passed],
        ["Jobs Failed", jobs_failed],
    ]
    
    for row in summary_data:
        ws1.append(row)
        
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 15

    # --- Sheet 2: Test Details ---
    ws2 = out_wb.create_sheet("Test Details")
    headers2 = ["Test / Job", "Status", "Total Tests", "Passed", "Failed", "Skipped", "Errors", "Pass %", "Fail %", "Duration (s)"]
    ws2.append(headers2)
    
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for job, m in job_metrics.items():
        job_pass_pct = (m["Passed"] / m["Total"] * 100) if m["Total"] > 0 else 0.0
        job_fail_pct = (m["Failed"] / m["Total"] * 100) if m["Total"] > 0 else 0.0
        status = "PASSED" if (m["Failed"] == 0 and m["Errors"] == 0) else "FAILED"
        
        row = [
            job, status, m["Total"], m["Passed"], m["Failed"], m["Skipped"], m["Errors"],
            f"{job_pass_pct:.2f}%", f"{job_fail_pct:.2f}%", f"{m['Duration']:.2f}"
        ]
        ws2.append(row)
        
        # Color status
        status_cell = ws2.cell(row=ws2.max_row, column=2)
        status_cell.font = Font(color="00B050" if status == "PASSED" else "FF0000", bold=True)
        
    for col in "ABCDEFGHIJ":
        ws2.column_dimensions[col].width = 15
    ws2.column_dimensions['A'].width = 30

    # --- Sheet 3: Failed Tests ---
    ws3 = out_wb.create_sheet("Failed Tests")
    headers3 = ["Test / Job", "Test Name", "Duration (s)", "Error Message"]
    ws3.append(headers3)
    
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
        
    if not failed_tests:
        ws3.append(["No failed tests!", "", "", ""])
    else:
        for f in failed_tests:
            ws3.append([f["Job"], f["Test"], f"{f['Duration']:.2f}", f["Error"]])
            
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 80

    os.makedirs('Test Results/Summary', exist_ok=True)
    out_file = "Test Results/Summary/Final_Test_Summary.xlsx"
    out_wb.save(out_file)
    print(f"Successfully generated final summary at {out_file}")

if __name__ == "__main__":
    generate_summary()
